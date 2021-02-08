from datetime import datetime

# from icecream import ic
import openpyxl





# Global variables
staff_details = {}  # contains permanent staff information.
MED_BILL_FILE = ''
STAFF_LIST_FILE = ''
UNDO_ENTRY_HISTORY = []  # list of entries to be undone.
REDO_ENTRY_HISTORY = []  # list of undone entries to be redone.


class MBillsFunctions:
    """
    Class which contains 'back-end' operations to be carried out on medical bills files.
    """

    @staticmethod
    def initializeFiles(med_bill_file=None, staff_list_file=None):
        """
        Loads Med Bills File and Staff List File into workbook objects.

        :param med_bill_file: Medical Bills File.

        :param staff_list_file: Staff List File.

        :return: Both workbooks, one of them, or None if one is of type: None.
        """
        global MED_BILL_FILE, STAFF_LIST_FILE
        if med_bill_file and staff_list_file is not None:
            workbook1 = openpyxl.load_workbook(med_bill_file)
            workbook2 = openpyxl.load_workbook(staff_list_file, read_only=True)
            MED_BILL_FILE, STAFF_LIST_FILE = med_bill_file, staff_list_file
            return workbook1, workbook2
        elif med_bill_file is None:
            workbook = openpyxl.load_workbook(staff_list_file, read_only=True)
            STAFF_LIST_FILE = staff_list_file
            return workbook
        elif staff_list_file is None:
            workbook = openpyxl.load_workbook(med_bill_file)
            MED_BILL_FILE = med_bill_file
            return workbook
        else:
            return None


    @staticmethod
    def getAllMedBillsNamesAndDept(workbook):
        """
        Static method to get all names from Med Bills workbook.

        :param workbook: Medical Bills workbook.

        :return: list of all people in the Medical Bills workbook.
        """
        people = []

        for sheet in workbook.sheetnames:
            for col in workbook[sheet].iter_cols(min_row=1, max_row=500, min_col=1, max_col=1):
                for cell in col:
                    # check for gray color, bold font, and whether cell is filled (not containing '0')
                    if cell.fill.start_color.index == 'FFD8D8D8' and cell.font.b is True \
                            and cell.value != 0:
                        people.append(cell.value.title() + '|' + workbook[sheet].title)

        return people


    @staticmethod
    def getAllDependantNames(workbook):  # todo: change name as it is misleading
        """
        Static method to extract all spouse, and children's names from Staff List workbook.

        :param workbook: Staff List Workbook.

        :return: list of all dependants from Staff list workbook.
        """

        d_names = []

        for sheet in workbook.sheetnames:
            for row in workbook[sheet].iter_rows(min_row=3, max_row=600, min_col=3, max_col=4):
                for cell in row:
                    if cell.value is not None and cell.value.isupper() is True:  # check for upper case letters and not empty cell
                        d_names.append(cell.value.title())
        return d_names


    @staticmethod
    def getDetailsOfPermanentStaff(workbook):
        """
        Optimized static method to extract details of permanent staff from the Staff List workbook.

        :param workbook: Staff List workbook.

        :return: dictionary of staff names->(Keys), their spouse and children->(Values).
        """
        sheet = workbook.active
        # start = datetime.now()

        def row_any(row):
            """
            Reimplementation of python's any() function but for rows.

            :param row: Row to be checked.

            :return: Returns False if all cells in a row are empty. True otherwise.
            """
            non_empty_cells = [x for x in row if x.value is not None]
            return any(non_empty_cells)

        helper_staff_name = ''  # to help link multiple children back to the original staff.

        def process_row(_row):
            """
            Function to process rows and create a dictionary of the staff's details.

            :param _row: Rows from Staff List workbook.

            :return: Dictionary containing all permanent staff's details.
            """
            global helper_staff_name, staff_details
            if _row[0].value is not None:
                helper_staff_name = _row[0].value

            # Assigning dept, spouse and child to staff name if staff is not empty
            # (meaning staff has multiple children according to the format of the file)
            if _row[0].value is not None:
                staff_details[_row[0].value] = [cell for cell in [_row[1].value, _row[2].value, _row[3].value]]
            else:
                if helper_staff_name == '':
                    raise Exception('Staff name provided was None!!')
                else:
                    staff_details[helper_staff_name].append(_row[-1].value)

            return staff_details

        # Filtering out rows which are empty which is ~40% of all rows...improving performance
        rows = [x for x in sheet.iter_rows(min_row=3, max_row=600, min_col=1, max_col=4) if row_any(x)]
        for row in rows:
            process_row(row)

        # stop = datetime.now()
        # ic('Time elapsed for extracting:', (stop - start))
        return staff_details


    @staticmethod
    def searchForStaffFromStaffList(person, staff_deets):
        """
        Static method to search for anyone using the Staff List workbook. If found, returns staff's details.
        If not, returns None for those particulars in the details(For Casuals and Guests).

        :param staff_deets: Dictionary of staff's details.

        :param person: Person to search for.

        :return: Tuple of staff with dependant(s) or None if not found.
        """
        # start = datetime.now()
        for staff, dependants in staff_deets.items():
            # ic.enable()
            if staff == person:
                # ic('Found with key:', staff, dependants)
                # stop = datetime.now()
                # ic('Time for Search elapsed:', stop - start)
                return staff.title(), dependants, 'k'
            else:
                for d in dependants:
                    if d == person:
                        # ic('Found with value:', staff, dependants)
                        # stop = datetime.now()
                        # ic('Time for Search elapsed:', stop - start)
                        return staff, dependants, 'v'
        return None, None, None


    @staticmethod
    def searchForCasualOrGuest(people_in_med_bill, person):
        """
        Static method to search specifically for a Guest or Casual since they are not in the Staff List workbook.

        :param people_in_med_bill: List of all people in Med Bills workbook with departments.

        :param person: Person to be searched for.

        :return: Name of Casual or Guest being searched for.
        """
        temp = [p for p in people_in_med_bill if p.split('|')[0] == person]
        return temp[0] if temp != [] else None


    @staticmethod
    def getDepartmentFromName(person: str, all_people_and_dept: list):
        """
        Static method that scans Med Bills workbook for a person's department.

        :param all_people_and_dept: List of all people in Med Bills workbook.

        :param person: Person to be searched for.

        :return: Returns Department of the person passed in.
        """
        for names in all_people_and_dept:
            if person == names.split('|')[0]:
                # ic.disable()
                # ic(names.split('|')[1])
                return names.split('|')[1]
        return None


    @staticmethod
    def getPersonAmountForMonth(workbook, person: str, all_people: list, months: dict, month: str):
        """
        Static method to get the current amount for the month of a person in the Med Bills workbook.

        :param all_people: List of everyone in Medical Bills workbook.

        :param workbook: Med Bills workbook.

        :param month: Specific month to extract amount from (key from months dict).

        :param months: Dictionary with months(keys) and offsets(values).

        :param person: Name of Person in Med Bills workbook.

        :return: Amount from cell.
        """
        # s1 = datetime.now()
        dept = MBillsFunctions.getDepartmentFromName(person, all_people)

        def processCellValue(celll):  # dont want problems with 'cell' from outer scope
            """
            Helper function for processing cell value to display in app.

            :param celll: Cell to perform operation on.

            :return: String of cell value appropriate for display.
            """
            if '=' in str(celll.value):
                temp = str(celll.value)[1:]
                if '+' in temp:
                    digits = temp.split('+')
                    temp = sum([float(x) for x in digits])
                amt = float(temp)
                return f'{amt:.2f}'
            else:
                return f'{celll.value:.2f}'  # returns 0.00

        sheet = workbook[dept]
        for col in sheet.iter_cols(min_row=4, max_row=500, min_col=1, max_col=1):
            for cell in col:
                if cell.value == person:
                    staff_cell = cell.offset(row=0, column=months.get(month, 0))
                    staff_amt = processCellValue(staff_cell)
                    child_cell = cell.offset(row=1, column=months.get(month, 0))
                    child_amt = processCellValue(child_cell)
                    spouse_cell = cell.offset(row=2, column=months.get(month, 0))
                    spouse_amt = processCellValue(spouse_cell)

                    # s2 = datetime.now()
                    # ic.enable()
                    # ic('Time taken to get amount:', s2 - s1)
                    # ic('Amounts:', staff_amt, child_amt, spouse_amt)
                    return staff_amt, child_amt, spouse_amt


    @staticmethod
    def insertAmountIntoMedBills(workbook, person: str, dept: str, offset_col: int, offset_row: int, amount: str):
        """
        Static method to insert amount into specific month of staff in Med Bills workbook.

        :param workbook: Medical Bills workbook.

        :param person: Staff/Guest/Casual.

        :param dept: Department of Staff.

        :param offset_col: Offset for month of entry.

        :param offset_row: Offset for staff/spouse/child.

        :param amount: Amount to be entered.

        :return: Boolean on whether operation was successful or not.
        """
        # start = datetime.now()
        # ic.enable()
        global UNDO_ENTRY_HISTORY
        wb = workbook
        sheet = wb[dept]

        for row in sheet.iter_rows(min_row=4, max_row=500, min_col=1, max_col=1):
            for cell in row:
                if cell.value == person:
                    c2 = cell.offset(row=offset_row, column=offset_col)
                    UNDO_ENTRY_HISTORY.append([wb, sheet, c2])
                    if c2.value == 0:
                        c2.value = '=' + str(amount)
                        # stop = datetime.now()
                        # ic('Time for actual insertion:', stop - start)
                        # ic('Amount inserted:', amount)
                        return True
                    else:
                        c2.value = str(c2.value) + '+' + str(amount)
                        # stop = datetime.now()
                        # ic('Time for actual insertion:', stop - start)
                        # ic('Amount inserted:', amount)
                        return True

        return False


    @staticmethod
    def undoEntry():
        """
        Static method to undo an entry.

        :return: Boolean indicating whether operation was successful or not.
        """
        global UNDO_ENTRY_HISTORY, REDO_ENTRY_HISTORY
        # start = datetime.now()
        # ic.enable()
        last_row_data = UNDO_ENTRY_HISTORY.pop()  # last set of values removed
        # REDO_ENTRY_HISTORY.clear()  # always clear to make sure only one value is stored
        REDO_ENTRY_HISTORY = last_row_data
        wb, sheet, cell = last_row_data[0], last_row_data[1], last_row_data[-1]
        a_cell = sheet.cell(row=cell.row, column=cell.column)

        if a_cell.value == 0:
            raise Exception('Cell is already at default value!')
        elif ('=' and '+') in a_cell.value:  # multiple amounts entered
            total_amount = a_cell.value.split('+')
            last_amount = total_amount.pop()
            # print('Last amount:', last_amount)
            rest_of_amount = '+'.join(total_amount)
            a_cell.value = rest_of_amount
            REDO_ENTRY_HISTORY.append(last_amount)
            # print('Rest of amount:', rest_of_amount)
            # stop = datetime.now()
            # ic('Time taken for undo:', stop-start)
            return True
        elif '=' in a_cell.value:  # just one amount entered
            cur_amount = a_cell.value
            # print('Cur amount:', cur_amount)
            a_cell.value = 0
            REDO_ENTRY_HISTORY.append(cur_amount)
            # stop = datetime.now()
            # ic('Time taken for undo:', stop - start)
            return True

        return False


    @staticmethod
    def redoEntry():
        """
        Static method to redo a previously undone entry.

        :return: Boolean value indicating success status.
        """
        global REDO_ENTRY_HISTORY
        last_undo_row = REDO_ENTRY_HISTORY  # will get overwritten so no need to pop but whatever
        wb, sheet, cell, amount = last_undo_row[0], last_undo_row[1], last_undo_row[2], last_undo_row[-1]
        a_cell = sheet.cell(row=cell.row, column=cell.column)

        if a_cell.value == 0:
            a_cell.value = amount
            return True
        elif '=' in a_cell.value:
            a_cell.value = a_cell.value + '+' + amount
            return True

        return False


    @staticmethod
    def saveFile(workbook):
        """
        Static method to save a workbook provided.

        :param workbook: Medical Bills workbook.

        :return: Boolean indicating whether save was successful or not.
        """
        # start = datetime.now()
        try:
            workbook.save(MED_BILL_FILE)
        except PermissionError:
            raise Exception('Couldn\'t write to the file because it is probably open??')
        except:
            raise Exception('There was a problem saving!')
        else:
            # stop = datetime.now()
            # ic.enable()
            # ic('Time taken to save:', stop-start)
            return True



    # ---------------------------------------- TODO --------------------------------------------------------
    # TODO:
